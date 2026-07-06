"""Mock-free, LLM-free structural self-test for the Meridian fixture."""
from __future__ import annotations
import json, sqlite3
from collections import Counter
from pathlib import Path
import fitz, yaml
from openpyxl import load_workbook
from proofbench.models import AuditConfig, Claim

ROOT=Path(__file__).resolve().parents[1]; AID="audit-2026-q1-meridian"; BASE=ROOT/"audits"/AID
gold=yaml.safe_load((BASE/"gold.yaml").read_text()); assert len(gold["claims"])==28
assert Counter(c["expected_status"] for c in gold["claims"])==Counter(supported=17,contradicted=3,ambiguous=3,outdated=3,missing_evidence=2)
cfg=yaml.safe_load((BASE/"audit.yaml").read_text()); assert len(cfg["documents"])==9; AuditConfig.model_validate(cfg)
for path in sorted((BASE/"claims").glob("claim-*.json")): Claim.model_validate(json.loads(path.read_text()))
assert sum(c["tolerance_rule"]["kind"]=="formula" for c in gold["claims"])==4
master=fitz.open(ROOT/cfg["documents"][0]["path"]); assert len(master)==4 and all(len(p.get_text())>300 for p in master)
for ref in cfg["documents"]:
    if ref["format"]=="pdf":
        d=fitz.open(ROOT/ref["path"]); assert all(len(p.get_text())>150 for p in d)
        if ref["kind"]=="vault": assert sum(len(p.find_tables().tables) for p in d)>=1, ref["doc_id"]
db=sqlite3.connect(ROOT/"index"/"search"/f"{AID}.db")
entities=db.execute("select count(*) from entities").fetchone()[0]; assert 60<=entities<=100,entities
names={r[0] for r in db.execute("select name from entities")}
for expected in ["Net revenue","FTE, period end","Enterprise accounts","Net operating cash outflow","Scope 2 location based","Contract personnel"]: assert expected in names,expected
facts=db.execute("select count(*) from facts").fetchone()[0]; assert facts>=180,facts
near=db.execute("select count(*) from edges e join entities n on n.entity_id=e.target_entity_id where n.name='Near coincidence control'").fetchone()[0]; assert near==0,near
totals=db.execute("select count(*) from facts where entity='Total'").fetchone()[0]; assert totals>=6,totals
headerless=db.execute("select count(*) from facts where entity in ('Office sites','Certified suppliers')").fetchone()[0]; assert headerless>=2,headerless
customer=load_workbook(BASE/"vault/customer-metrics.xlsx",data_only=False)["Metrics"]
formats={cell.value: customer.cell(cell.row,2).number_format for cell in customer["A"] if cell.value}
assert formats["Net dollar retention"]=="0.0%" and formats["Annualized subscription value"]=='€#,##0'
print(json.dumps({"claims":28,"documents":9,"entities":entities,"facts":facts,"near_coincidence_edges":near,"status":"PASS"},indent=2))
