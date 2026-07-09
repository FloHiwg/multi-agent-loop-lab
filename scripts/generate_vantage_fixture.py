"""Deterministically generate the Vantage fixture: a scaled vault benchmark.

Why it exists (PROTOCOL.md, exp-20260709T125305Z): at Meridian's 8-doc scale,
baseline FTS and the graph tools reach the same accuracy, the researcher
sub-agent goes unused, and the graph's scaling argument stays untested. This
fixture is the bigger world: ~40 vault documents over 5 quarters, and --
new vs Meridian -- claims whose evidence lives in narrative BODY TEXT, not
tables. Prose spans are indexed page-level in spans_fts but never enter the
facts graph, so narrative claims exercise exactly what entity_profile
cannot see.

Design stance (same as Meridian, scaled):
- A seeded world model owns every number; documents are renderings of it;
  gold labels derive mechanically from the perturbations the generator
  itself plants. No LLM anywhere, byte-identical regeneration.
- Perturbations are parameterized per failure class, and every gold entry
  carries a `failure_class` tag so evals can slice by class.

Failure classes planted:
  table_vocab        supported, table evidence, synonym wording
  narrative_only     supported, evidence exists ONLY in prose
  formula            supported, derived figure (margin, per-FTE)
  scale_trap         supported, kEUR source vs EUR-million claim
  contradicted_table contradicted by a table value
  contradicted_prose contradicted by a prose-only value
  prose_table_clash  ambiguous: quarterly update prose disagrees with the
                     finance table, no stated ordering
  regional_clash     ambiguous: regional cut vs finance, explicitly unranked
  superseded_v1      outdated: matches preliminary v1, restated by v2
  stale_quarter      outdated: asserts the PRIOR quarter's value as current
  absent_near_name   missing_evidence with a near-name distractor present
"""
from __future__ import annotations

import argparse
import json
import random
import re
import zipfile
from datetime import datetime
from pathlib import Path

import fitz
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
AUDIT_ID = "audit-2026-q2-vantage"
BASE = ROOT / "audits" / AUDIT_ID
VAULT = BASE / "vault"
MASTER = BASE / "master" / "vantage-q2-exec-review.pdf"
COMPANY = "Vantage Industrial Group SE"

NAVY, RUST, PALE, INK, LINE = (0.16, 0.17, 0.24), (0.64, 0.31, 0.16), (0.96, 0.94, 0.90), (0.14, 0.15, 0.18), (0.80, 0.80, 0.78)

QUARTERS = ["2025-Q2", "2025-Q3", "2025-Q4", "2026-Q1", "2026-Q2"]
QLABEL = {q: f"Q{q[-1]} {q[:4]}" for q in QUARTERS}
REGIONS = ["DACH", "France", "Nordics", "UK and Ireland", "Iberia", "Benelux"]


# ---------------------------------------------------------------- world model

def build_world(seed: int) -> dict:
    """Every metric's value per quarter, in integer kEUR / counts / bp so all
    arithmetic is exact. Deterministic in (seed, definition order)."""
    rng = random.Random(seed)
    w: dict[str, dict[str, int]] = {}

    def series(name: str, start: int, lo: float, hi: float) -> None:
        vals, v = {}, start
        for q in QUARTERS:
            vals[q] = v
            v = int(round(v * (1 + rng.uniform(lo, hi))))
        w[name] = vals

    # regional revenue (kEUR) -- revenue is their exact sum, so the
    # aggregation edge mines and regional tables foot
    for region, start in zip(REGIONS, (6200, 3900, 2600, 4400, 1700, 1400)):
        series(f"region:{region}", start, 0.005, 0.045)
    w["revenue"] = {q: sum(w[f"region:{r}"][q] for r in REGIONS) for q in QUARTERS}
    series("cost_of_delivery", 8300, 0.004, 0.035)
    w["gross_profit"] = {q: w["revenue"][q] - w["cost_of_delivery"][q] for q in QUARTERS}
    series("opex", 6900, 0.002, 0.030)
    w["operating_result"] = {q: w["gross_profit"][q] - w["opex"][q] for q in QUARTERS}
    series("cash_operating", 4100, -0.02, 0.05)
    series("cash_deposit", 2600, -0.01, 0.04)
    w["cash_total"] = {q: w["cash_operating"][q] + w["cash_deposit"][q] for q in QUARTERS}
    series("net_op_outflow", 1450, -0.05, 0.06)     # sign rendered as (…)
    series("pipeline_fin", 11800, 0.01, 0.06)
    w["pipeline_reg"] = {q: w["pipeline_fin"][q] - rng.randint(280, 520) for q in QUARTERS}
    series("deferred_revenue", 9100, 0.005, 0.03)
    series("receivables", 6300, 0.0, 0.035)
    series("asv", 30400, 0.01, 0.05)                # annualized subscription value
    series("expansion_bookings", 3600, 0.0, 0.06)

    series("fte", 213, 0.0, 0.03)
    series("contractors", 26, -0.03, 0.05)
    series("enterprise_accounts", 302, 0.005, 0.03)
    w["new_wins_ops"] = {q: rng.randint(38, 62) for q in QUARTERS}
    w["new_wins_cust"] = {q: w["new_wins_ops"][q] - rng.choice((2, 3, 4)) for q in QUARTERS}
    series("licensed_seats", 9450, 0.01, 0.04)
    series("open_cases", 410, -0.05, 0.05)
    series("learning_hours", 1650, -0.04, 0.06)
    series("scope2", 585, -0.04, 0.01)
    w["attrition_bp"] = {q: rng.randint(24, 41) for q in QUARTERS}      # 0.1% units
    w["ndr_bp"] = {q: rng.randint(1060, 1140) for q in QUARTERS}
    w["sla_bp"] = {q: rng.randint(920, 965) for q in QUARTERS}

    # narrative-only metrics: appear in prose documents, never in any table
    w["nps"] = {q: rng.randint(38, 57) for q in QUARTERS}
    w["renewal_bp"] = {q: rng.randint(905, 955) for q in QUARTERS}
    w["onboarding_days"] = {q: rng.randint(24, 39) for q in QUARTERS}
    w["certified_partners"] = {q: rng.randint(50, 90) for q in QUARTERS}
    w["prod_incidents"] = {q: rng.randint(1, 6) for q in QUARTERS}

    # planted deltas for the latest quarter's perturbations
    latest = QUARTERS[-1]
    w["_v1_revenue_delta"] = {latest: rng.randint(140, 260)}   # v1 = v2 - delta
    w["_update_pipeline_delta"] = {latest: rng.randint(180, 320)}
    return w


def fm(kEUR: int) -> str:
    return f"{kEUR / 1000:.3f}"          # table style: EUR millions, 3dp


def fk(kEUR: int) -> str:
    s = f"{abs(kEUR):,}"
    return f"({s})" if kEUR < 0 else s   # treasury style: kEUR with parens


def fpct(bp: int) -> str:
    return f"{bp / 10:.1f}%"


def prose_m(kEUR: int) -> str:
    return f"EUR {kEUR / 1000:.2f} million"


# ---------------------------------------------------------------- rendering

def pdf_header(page, title, subtitle):
    page.insert_text((50, 48), "VANTAGE INDUSTRIAL GROUP", fontsize=9, fontname="hebo", color=RUST)
    page.insert_text((50, 78), title, fontsize=21, fontname="hebo", color=NAVY)
    page.insert_text((50, 98), subtitle, fontsize=10, color=(0.42, 0.42, 0.40))
    page.draw_line((50, 112), (545, 112), color=LINE, width=.7)


def pdf_footer(page, n):
    page.draw_line((50, 805), (545, 805), color=LINE, width=.6)
    page.insert_text((50, 820), "Confidential | Synthetic Proofbench fixture", fontsize=7, color=(.42, .42, .40))
    page.insert_text((510, 820), f"Page {n}", fontsize=7, color=(.42, .42, .40))


def draw_table(page, y, rows, widths=None, font=7.2):
    cols = max(len(r) for r in rows)
    widths = widths or [495 / cols] * cols
    h = 22
    for ri, row in enumerate(rows):
        x = 50
        for ci in range(cols):
            w = widths[ci]
            rect = fitz.Rect(x, y + ri * h, x + w, y + (ri + 1) * h)
            page.draw_rect(rect, color=LINE, fill=NAVY if ri == 0 else (PALE if ri % 2 == 0 else (1, 1, 1)), width=.5)
            text = str(row[ci]) if ci < len(row) else ""
            page.insert_textbox(rect + (5, 5, -4, -3), text, fontsize=font,
                                fontname="hebo" if ri == 0 else "helv",
                                color=(1, 1, 1) if ri == 0 else INK, align=0)
            x += w
    return y + len(rows) * h


def save_pdf(doc, path):
    doc.set_metadata({"title": path.stem, "author": COMPANY,
                      "creationDate": "D:20260715090000+02'00'", "modDate": "D:20260715090000+02'00'"})
    path.parent.mkdir(parents=True, exist_ok=True)
    doc.save(path, garbage=4, deflate=True, clean=True, no_new_id=True)


def make_pdf(path, title, pages):
    """pages: list of (subtitle, tables, paragraphs); tables may be empty for
    prose-only documents. Paragraphs render as flowing body text."""
    doc = fitz.open()
    for i, (subtitle, tables, paragraphs) in enumerate(pages, 1):
        p = doc.new_page(width=595, height=842)
        pdf_header(p, title, subtitle)
        y = 135
        for heading, rows, widths in tables:
            p.insert_text((50, y), heading, fontsize=13, fontname="hebo", color=NAVY)
            y += 14
            y = draw_table(p, y, rows, widths) + 22
        for para in paragraphs:
            box = fitz.Rect(50, y, 545, y + 120)
            p.insert_textbox(box, para, fontsize=9.5, lineheight=1.45, color=INK)
            y += 30 + 13 * (len(para) // 95 + 1)
        pdf_footer(p, i)
    save_pdf(doc, path)


def normalize_xlsx(path):
    tmp = path.with_suffix(".normalized.xlsx")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as zout:
        for name in sorted(zin.namelist()):
            info = zipfile.ZipInfo(name, (1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            payload = zin.read(name)
            if name == "docProps/core.xml":
                payload = re.sub(rb"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                                 rb'<dcterms:modified xsi:type="dcterms:W3CDTF">2026-07-15T09:00:00Z</dcterms:modified>', payload)
            zout.writestr(info, payload)
    tmp.replace(path)


def make_xlsx(path, title, blocks):
    wb = Workbook()
    wb.remove(wb.active)
    fixed = datetime(2026, 7, 15, 9, 0, 0)
    wb.properties.creator = COMPANY
    wb.properties.lastModifiedBy = COMPANY
    wb.properties.created = fixed
    wb.properties.modified = fixed
    wb.properties.revision = "1"
    for sheet_name, block_sets in blocks:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:D1")
        ws["A1"] = title
        ws["A1"].fill = PatternFill("solid", fgColor="2A2B3D")
        ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
        row = 4
        for heading, rows in block_sets:
            ws.cell(row, 1, heading).font = Font(bold=True, color="2A2B3D", size=12)
            row += 1
            for values in rows:
                for col, value in enumerate(values, 1):
                    ws.cell(row, col, value)
                if values == rows[0]:
                    for cell in ws[row]:
                        cell.fill = PatternFill("solid", fgColor="A24F28")
                        cell.font = Font(color="FFFFFF", bold=True)
                row += 1
            row += 2
        for col, w in enumerate([34, 18, 18, 18], 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(vertical="center")
                cell.border = Border(bottom=Side(style="hair", color="DAD5CC"))
            label = str(row_cells[0].value or "").casefold()
            if any(word in label for word in ("retention", "attrition", "attainment")):
                for cell in row_cells[1:4]:
                    cell.number_format = "0.0%"
            elif any(word in label for word in ("value", "bookings")):
                for cell in row_cells[1:4]:
                    cell.number_format = '€#,##0'
            elif row_cells[0].value not in (None, "Metric"):
                for cell in row_cells[1:4]:
                    cell.number_format = "#,##0"
        ws.freeze_panes = "A4"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    normalize_xlsx(path)


# ---------------------------------------------------------------- documents

def doc_id(family: str, q: str) -> str:
    return f"vantage-{family}-{q.lower().replace('-', '')}"


def generate_documents(w: dict) -> list[tuple[str, str, str, str]]:
    """Render every vault document. Returns audit.yaml document rows."""
    docs: list[tuple[str, str, str, str]] = []
    latest = QUARTERS[-1]

    def hdr(q):
        prev = QUARTERS[QUARTERS.index(q) - 1] if QUARTERS.index(q) else None
        return ["EUR million", f"{QLABEL[q]} actual", f"{QLABEL[prev]} actual" if prev else "n/a", f"{QLABEL[q]} budget"]

    def cells(name, q, fmtf=fm, budget_shift=0.97):
        prev = QUARTERS[QUARTERS.index(q) - 1] if QUARTERS.index(q) else None
        return [fmtf(w[name][q]), fmtf(w[name][prev]) if prev else "-", fmtf(int(w[name][q] * budget_shift))]

    for q in QUARTERS:
        ql = QLABEL[q]
        prev = QUARTERS[QUARTERS.index(q) - 1] if QUARTERS.index(q) else None

        # -- finance pack (v1+v2 for the latest quarter, single pack before)
        def finance_rows(revenue_kEUR, gp_kEUR, opres_kEUR):
            return [hdr(q),
                    ["Net revenue"] + [fm(revenue_kEUR)] + cells("revenue", q)[1:],
                    ["Cost of delivery", f"({fm(w['cost_of_delivery'][q])})",
                     f"({fm(w['cost_of_delivery'][QUARTERS[QUARTERS.index(q)-1]])})" if QUARTERS.index(q) else "-",
                     f"({fm(int(w['cost_of_delivery'][q] * 0.97))})"],
                    ["Gross profit"] + [fm(gp_kEUR)] + cells("gross_profit", q)[1:],
                    ["Adjusted operating result"] + [fm(opres_kEUR)] + cells("operating_result", q)[1:],
                    ["Deferred revenue"] + cells("deferred_revenue", q),
                    ["Trade receivables"] + cells("receivables", q)]

        pipeline_rows = [hdr(q), ["Qualified pipeline"] + cells("pipeline_fin", q),
                         ["Expansion bookings"] + cells("expansion_bookings", q)]
        if q == latest:
            d = w["_v1_revenue_delta"][latest]
            rev1 = w["revenue"][q] - d
            make_pdf(VAULT / f"finance-pack-{q}-v1.pdf", f"{ql} Finance Pack v1",
                     [("Pre-close management accounts",
                       [("Income statement", finance_rows(rev1, w["gross_profit"][q] - d, w["operating_result"][q] - d), [150, 115, 115, 115]),
                        ("Commercial indicators", pipeline_rows, [150, 115, 115, 115])],
                       [f"PRELIMINARY - superseded on 15 July 2026 by restated v2."])])
            docs.append((doc_id("finance-v1", q), f"finance-pack-{q}-v1.pdf", "pdf", "finance_pack"))
            title, fname, note = f"{ql} Finance Pack v2", f"finance-pack-{q}-v2.pdf", "Final restatement, approved 15 July 2026."
            did = doc_id("finance-v2", q)
        else:
            title, fname, note = f"{ql} Finance Pack", f"finance-pack-{q}.pdf", f"Final as of quarter close {ql}."
            did = doc_id("finance", q)
        make_pdf(VAULT / fname, title,
                 [("Management accounts",
                   [("Income statement", finance_rows(w["revenue"][q], w["gross_profit"][q], w["operating_result"][q]), [150, 115, 115, 115]),
                    ("Commercial indicators", pipeline_rows, [150, 115, 115, 115])],
                   [note])])
        docs.append((did, fname, "pdf", "finance_pack"))

        # -- operations review
        ops = [["Metric", ql, "Prior quarter", "Plan"],
               ["FTE, period end", str(w["fte"][q]), str(w["fte"][QUARTERS[QUARTERS.index(q)-1]]) if QUARTERS.index(q) else "-", str(w["fte"][q] + 4)],
               ["P1 response SLA attainment", fpct(w["sla_bp"][q]), "-", "95.0%"],
               ["Open case inventory", str(w["open_cases"][q]), "-", str(int(w["open_cases"][q] * 0.9))],
               ["New enterprise accounts", str(w["new_wins_ops"][q]), "-", str(w["new_wins_ops"][q] - 3)],
               ["Production incidents", str(w["prod_incidents"][q]), "-", "0"]]
        make_pdf(VAULT / f"operations-review-{q}.pdf", f"{ql} Operations Review",
                 [("Service delivery", [("Operating metrics", ops, [205, 95, 95, 95])],
                   [f"All metrics final as of quarter close {ql}."])])
        docs.append((doc_id("operations", q), f"operations-review-{q}.pdf", "pdf", "operations_review"))

        # -- customer workbook
        pq = prev or q  # first quarter repeats itself as its own "prior"
        make_xlsx(VAULT / f"customer-metrics-{q}.xlsx", f"{ql} Customer Metrics", [("Metrics", [
            ("Customer base", [["Metric", ql, "Prior", "Plan"],
                               ["Enterprise accounts", w["enterprise_accounts"][q], w["enterprise_accounts"][pq], w["enterprise_accounts"][q] - 2],
                               ["New enterprise accounts", w["new_wins_cust"][q], w["new_wins_cust"][pq], w["new_wins_cust"][q] + 3],
                               ["Gross account attrition", w["attrition_bp"][q] / 1000, w["attrition_bp"][pq] / 1000, 0.03]]),
            ("Recurring value", [["Metric", ql, "Prior", "Plan"],
                                 ["Annualized subscription value", w["asv"][q] * 1000, w["asv"][pq] * 1000, (w["asv"][q] - 400) * 1000],
                                 ["Net dollar retention", w["ndr_bp"][q] / 1000, w["ndr_bp"][pq] / 1000, 1.1],
                                 ["Licensed seats", w["licensed_seats"][q], w["licensed_seats"][pq], w["licensed_seats"][q] + 200]])])])
        docs.append((doc_id("customer", q), f"customer-metrics-{q}.xlsx", "xlsx", "customer_appendix"))

        # -- regional performance (foots to revenue; unranked pipeline cut)
        reg = [["Region", f"{ql} actual", "Prior actual", f"{ql} budget"]]
        prev = QUARTERS[QUARTERS.index(q) - 1] if QUARTERS.index(q) else None
        for r in REGIONS:
            reg.append([r, fm(w[f"region:{r}"][q]), fm(w[f"region:{r}"][prev]) if prev else "-", fm(int(w[f"region:{r}"][q] * 0.97))])
        reg.append(["Total", fm(w["revenue"][q]), fm(w["revenue"][prev]) if prev else "-", fm(sum(int(w[f"region:{r}"][q] * 0.97) for r in REGIONS))])
        reg.append(["Qualified commercial pipeline", fm(w["pipeline_reg"][q]), "-", "-"])
        make_pdf(VAULT / f"regional-performance-{q}.pdf", f"{ql} Regional Performance",
                 [("Revenue by market", [("Regional scorecard", reg, [145, 115, 115, 115])],
                   ["Regional pipeline uses local CRM qualification and is not ranked against Finance."])])
        docs.append((doc_id("regional", q), f"regional-performance-{q}.pdf", "pdf", "regional_review"))

        # -- treasury (kEUR scale trap)
        cash = [["kEUR", f"{ql} close", "Prior close", "Plan"],
                ["Operating accounts", fk(w["cash_operating"][q]), "-", "-"],
                ["Deposit accounts", fk(w["cash_deposit"][q]), "-", "-"],
                ["Total", fk(w["cash_total"][q]), "-", "-"],
                ["Net operating cash outflow", fk(-w["net_op_outflow"][q]), "-", "-"]]
        make_pdf(VAULT / f"treasury-{q}.pdf", f"{ql} Treasury and Cash",
                 [("Cash position", [("Treasury ledger", cash, [185, 105, 105, 100])],
                   ["All figures in thousands of euros."])])
        docs.append((doc_id("treasury", q), f"treasury-{q}.pdf", "pdf", "treasury"))

        # -- people workbook
        make_xlsx(VAULT / f"people-review-{q}.xlsx", f"{ql} People Review", [("People", [
            ("Workforce", [["Metric", ql, "Prior", "Plan"],
                           ["Contract personnel", w["contractors"][q], w["contractors"][pq], w["contractors"][q] + 1],
                           ["Structured learning hours", w["learning_hours"][q], w["learning_hours"][pq], w["learning_hours"][q] + 100],
                           ["Scope 2 location based", w["scope2"][q], w["scope2"][pq], w["scope2"][q] - 10]])])])
        docs.append((doc_id("people", q), f"people-review-{q}.xlsx", "xlsx", "people_review"))

        # -- management commentary: PROSE ONLY. Carries narrative-only
        # metrics (NPS, renewal rate, onboarding, certified partners) that
        # exist in no table anywhere.
        commentary = [
            f"Trading in {ql} developed in line with the restated plan. Net revenue reached {prose_m(w['revenue'][q])}, "
            f"with the gross result at {prose_m(w['gross_profit'][q])} and an adjusted operating result of {prose_m(w['operating_result'][q])}.",
            f"Customer sentiment strengthened again: the group-wide net promoter score came in at {w['nps'][q]} for the quarter, "
            f"and the gross renewal rate held at {fpct(w['renewal_bp'][q])} across the installed base.",
            f"Delivery efficiency improved. The median onboarding cycle for new enterprise customers was {w['onboarding_days'][q]} days, "
            f"and the certified implementation partner network grew to {w['certified_partners'][q]} partners at quarter end.",
            f"The workforce closed at {w['fte'][q]} full-time equivalents, supported by {w['contractors'][q]} external contract staff.",
        ]
        make_pdf(VAULT / f"management-commentary-{q}.pdf", f"{ql} Management Commentary",
                 [("Letter to the supervisory board", [], commentary)])
        docs.append((doc_id("commentary", q), f"management-commentary-{q}.pdf", "pdf", "management_commentary"))

        # -- quarterly update: prose brief. For the latest quarter its
        # pipeline number deliberately disagrees with the finance table
        # (prose_table_clash), with no stated ordering.
        pipeline_update = w["pipeline_fin"][q] - (w["_update_pipeline_delta"][latest] if q == latest else 0)
        update = [
            f"{COMPANY} closed {ql} with revenue of {prose_m(w['revenue'][q])} and {w['new_wins_ops'][q]} new enterprise account wins.",
            f"The qualified pipeline stood at {prose_m(pipeline_update)} entering the next quarter.",
            f"Annualized subscription value reached {prose_m(w['asv'][q])}, with licensed seats at {w['licensed_seats'][q]:,}.",
        ]
        make_pdf(VAULT / f"quarterly-update-{q}.pdf", f"{ql} Quarterly Update",
                 [("Investor and staff brief", [], update)])
        docs.append((doc_id("update", q), f"quarterly-update-{q}.pdf", "pdf", "quarterly_update"))

    return docs


# ---------------------------------------------------------------- claims

def build_claims(w: dict) -> list[dict]:
    """Claims about the latest quarter, each tagged with its failure class.
    Values derive from the same world the documents rendered, so gold labels
    are correct by construction."""
    L = QUARTERS[-1]
    P = QUARTERS[-2]
    lm = QLABEL[L]
    fin2 = doc_id("finance-v2", L)
    fin1 = doc_id("finance-v1", L)
    v1d = w["_v1_revenue_delta"][L]
    c: list[dict] = []

    def add(text, value, unit, status, fclass, source_doc, span, rule=None, notes=""):
        c.append({"text": text, "value": value, "unit": unit, "status": status,
                  "failure_class": fclass, "source_doc": source_doc, "span": span,
                  "rule": rule or {"kind": "exact"}, "notes": notes})

    e3 = lambda k: round(k * 1000)  # kEUR -> EUR

    # table_vocab: synonyms against table labels
    add(f"Group turnover for the quarter was {prose_m(w['revenue'][L])}.", e3(w["revenue"][L]), "currency",
        "supported", "table_vocab", fin2, f"Net revenue | {fm(w['revenue'][L])}", notes="turnover vs Net revenue")
    add(f"Operating contribution reached {prose_m(w['operating_result'][L])}.", e3(w["operating_result"][L]), "currency",
        "supported", "table_vocab", fin2, f"Adjusted operating result | {fm(w['operating_result'][L])}")
    add(f"Headcount at quarter end was {w['fte'][L]}.", w["fte"][L], "count",
        "supported", "table_vocab", doc_id("operations", L), f"FTE, period end | {w['fte'][L]}")
    add(f"The enterprise client base closed at {w['enterprise_accounts'][L]}.", w["enterprise_accounts"][L], "count",
        "supported", "table_vocab", doc_id("customer", L), f"Enterprise accounts | {w['enterprise_accounts'][L]}")
    add(f"Recurring revenue retention was {fpct(w['ndr_bp'][L])}.", w["ndr_bp"][L] / 1000, "percent",
        "supported", "table_vocab", doc_id("customer", L), f"Net dollar retention")
    add(f"Unearned revenue on the balance sheet was {prose_m(w['deferred_revenue'][L])}.", e3(w["deferred_revenue"][L]), "currency",
        "supported", "table_vocab", fin2, f"Deferred revenue | {fm(w['deferred_revenue'][L])}")
    add(f"DACH contributed {prose_m(w['region:DACH'][L])} of quarterly sales.", e3(w["region:DACH"][L]), "currency",
        "supported", "table_vocab", doc_id("regional", L), f"DACH | {fm(w['region:DACH'][L])}")
    add(f"The external delivery workforce numbered {w['contractors'][L]}.", w["contractors"][L], "count",
        "supported", "table_vocab", doc_id("people", L), f"Contract personnel | {w['contractors'][L]}")

    # narrative_only: evidence exists ONLY in prose documents
    add(f"The group net promoter score was {w['nps'][L]}.", w["nps"][L], "score",
        "supported", "narrative_only", doc_id("commentary", L), f"net promoter score came in at {w['nps'][L]}")
    add(f"Gross renewal rate held at {fpct(w['renewal_bp'][L])}.", w["renewal_bp"][L] / 1000, "percent",
        "supported", "narrative_only", doc_id("commentary", L), f"gross renewal rate held at {fpct(w['renewal_bp'][L])}")
    add(f"Median onboarding cycle for new enterprise customers was {w['onboarding_days'][L]} days.", w["onboarding_days"][L], "days",
        "supported", "narrative_only", doc_id("commentary", L), f"median onboarding cycle")
    add(f"The certified implementation partner network stood at {w['certified_partners'][L]} partners.", w["certified_partners"][L], "count",
        "supported", "narrative_only", doc_id("commentary", L), f"certified implementation partner network grew to {w['certified_partners'][L]}")

    # formula: derived figures
    margin = w["gross_profit"][L] / w["revenue"][L]
    add(f"The gross margin was {margin * 100:.1f}%.", round(margin, 4), "percent",
        "supported", "formula", fin2, f"Gross profit | {fm(w['gross_profit'][L])}",
        rule={"kind": "formula", "formula": "gross_profit / net_revenue", "tolerance_pct": 0.2})
    rev_per_fte = round(w["revenue"][L] * 1000 / w["fte"][L], 2)
    add(f"Quarterly sales per employee were about EUR {rev_per_fte / 1000:.1f} thousand.", rev_per_fte, "currency_per_fte",
        "supported", "formula", fin2, f"Net revenue | {fm(w['revenue'][L])}",
        rule={"kind": "formula", "formula": "net_revenue / fte_period_end", "tolerance_pct": 0.2},
        notes="cross-document: finance revenue / operations FTE")

    # scale_trap: kEUR treasury vs EUR-million claims
    add(f"Available liquidity at quarter end was {prose_m(w['cash_total'][L])}.", e3(w["cash_total"][L]), "currency",
        "supported", "scale_trap", doc_id("treasury", L), f"Total | {fk(w['cash_total'][L])}", notes="source is kEUR")
    add(f"Cash burn from operations was {prose_m(w['net_op_outflow'][L])}.", -e3(w["net_op_outflow"][L]), "currency",
        "supported", "scale_trap", doc_id("treasury", L), f"Net operating cash outflow | {fk(-w['net_op_outflow'][L])}",
        notes="kEUR and parentheses sign")

    # contradicted_table
    add(f"The closing workforce was {w['fte'][L] + 2} FTE.", w["fte"][L] + 2, "count",
        "contradicted", "contradicted_table", doc_id("operations", L), f"FTE, period end | {w['fte'][L]}", notes="off by two")
    add(f"The unresolved support queue ended at {w['open_cases'][L] - 12} cases.", w["open_cases"][L] - 12, "count",
        "contradicted", "contradicted_table", doc_id("operations", L), f"Open case inventory | {w['open_cases'][L]}")
    add(f"Licensed seats totalled {w['licensed_seats'][L] + 150:,}.", w["licensed_seats"][L] + 150, "count",
        "contradicted", "contradicted_table", doc_id("customer", L), f"Licensed seats | {w['licensed_seats'][L]}")

    # contradicted_prose: only prose evidence exists, and it disagrees
    add(f"The group net promoter score improved to {w['nps'][L] + 6}.", w["nps"][L] + 6, "score",
        "contradicted", "contradicted_prose", doc_id("commentary", L), f"net promoter score came in at {w['nps'][L]}")
    add(f"Median onboarding cycle was {w['onboarding_days'][L] - 5} days.", w["onboarding_days"][L] - 5, "days",
        "contradicted", "contradicted_prose", doc_id("commentary", L), f"median onboarding cycle")

    # prose_table_clash: quarterly update prose vs finance table, no ordering
    upd = w["pipeline_fin"][L] - w["_update_pipeline_delta"][L]
    add(f"Qualified pipeline stood at {prose_m(w['pipeline_fin'][L])}.", e3(w["pipeline_fin"][L]), "currency",
        "ambiguous", "prose_table_clash", fin2, f"Qualified pipeline | {fm(w['pipeline_fin'][L])}",
        notes=f"quarterly update prose says {prose_m(upd)}; no ordering stated")

    # regional_clash: regional cut vs finance, explicitly unranked
    add(f"The commercial pipeline was {prose_m(w['pipeline_reg'][L])}.", e3(w["pipeline_reg"][L]), "currency",
        "ambiguous", "regional_clash", doc_id("regional", L), f"Qualified commercial pipeline | {fm(w['pipeline_reg'][L])}",
        notes="finance pack reports a different value; regional explicitly unranked")
    add(f"New enterprise wins totalled {w['new_wins_ops'][L]}.", w["new_wins_ops"][L], "count",
        "ambiguous", "regional_clash", doc_id("operations", L), f"New enterprise accounts | {w['new_wins_ops'][L]}",
        notes=f"customer workbook says {w['new_wins_cust'][L]}; equal authority")

    # superseded_v1: matches preliminary v1 figures only
    add(f"Quarterly revenue came to {prose_m(w['revenue'][L] - v1d)}.", e3(w["revenue"][L] - v1d), "currency",
        "outdated", "superseded_v1", fin1, f"Net revenue | {fm(w['revenue'][L] - v1d)}")
    add(f"Gross profit was {prose_m(w['gross_profit'][L] - v1d)}.", e3(w["gross_profit"][L] - v1d), "currency",
        "outdated", "superseded_v1", fin1, f"Gross profit | {fm(w['gross_profit'][L] - v1d)}")

    # stale_quarter: prior quarter's value asserted as current
    add(f"Annualized subscription value stood at {prose_m(w['asv'][P])}.", e3(w["asv"][P]), "currency",
        "outdated", "stale_quarter", doc_id("customer", P), "Annualized subscription value",
        notes=f"that is the {QLABEL[P]} figure; {lm} differs")
    add(f"Structured learning hours totalled {w['learning_hours'][P]:,}.", w["learning_hours"][P], "hours",
        "outdated", "stale_quarter", doc_id("people", P), "Structured learning hours",
        notes=f"prior-quarter figure asserted as current")

    # absent_near_name: no evidence, near-name distractor exists
    add("Monthly active platform seats were 8,912.", 8912, "count",
        "missing_evidence", "absent_near_name", None, None,
        notes="Licensed seats exists as near-name distractor; monthly active seats appears nowhere")
    add("Weighted services backlog was EUR 5.30 million.", 5300000, "currency",
        "missing_evidence", "absent_near_name", None, None,
        notes="pipeline metrics exist as near-name distractors; backlog appears nowhere")
    add("Employee engagement index was 7.9.", 7.9, "score",
        "missing_evidence", "absent_near_name", None, None,
        notes="NPS exists in prose as a near-name score distractor")

    return c


def make_master(claims: list[dict]):
    doc = fitz.open()
    per_page = 7
    groups = [claims[i:i + per_page] for i in range(0, len(claims), per_page)]
    for pno, group in enumerate(groups, 1):
        p = doc.new_page(width=595, height=842)
        pdf_header(p, "Q2 2026 Executive Review", f"Key figures, part {pno}")
        y = 145
        for i, cl in enumerate(group):
            n = (pno - 1) * per_page + i + 1
            p.insert_textbox(fitz.Rect(58, y, 535, y + 48), f"{n:02d}. {cl['text']}",
                             fontsize=10.5, lineheight=1.35, color=INK)
            y += 58
        pdf_footer(p, pno)
    save_pdf(doc, MASTER)


def emit_metadata(claims: list[dict], docs: list[tuple[str, str, str, str]]):
    documents = [{"doc_id": "vantage-master-q2-2026", "path": str(MASTER.relative_to(ROOT)),
                  "kind": "master", "format": "pdf", "tag": None}]
    documents += [{"doc_id": d, "path": str((VAULT / f).relative_to(ROOT)), "kind": "vault",
                   "format": fmt, "tag": tag} for d, f, fmt, tag in docs]
    audit = {"audit_id": AUDIT_ID, "master_doc_id": "vantage-master-q2-2026",
             "default_tolerance_rule": {"kind": "exact"},
             "evidence_priority": ["finance_pack", "operations_review", "customer_appendix",
                                   "regional_review", "treasury", "people_review",
                                   "management_commentary", "quarterly_update"],
             "documents": documents, "created_at": "2026-07-15T09:00:00+02:00"}
    (BASE / "audit.yaml").write_text(yaml.safe_dump(audit, sort_keys=False, allow_unicode=False))

    gold = []
    (BASE / "claims").mkdir(parents=True, exist_ok=True)
    for n, cl in enumerate(claims, 1):
        cid = f"{AUDIT_ID}/claim-{n:04d}"
        gold.append({"claim_id": cid, "claim_text": cl["text"], "canonical_value": cl["value"],
                     "unit": cl["unit"], "currency": "EUR" if cl["unit"].startswith("currency") else None,
                     "expected_status": cl["status"], "failure_class": cl["failure_class"],
                     "source_doc": cl["source_doc"], "source_span": cl["span"],
                     "tolerance_rule": cl["rule"], "notes": cl["notes"]})
        claim = {"claim_id": cid, "label": cl["text"].rstrip("."), "raw_text": cl["text"],
                 "canonical_value": cl["value"], "unit": cl["unit"],
                 "currency": "EUR" if cl["unit"].startswith("currency") else None,
                 "entity": COMPANY, "time_scope": "Q2 2026", "tolerance_rule": cl["rule"],
                 "expected_evidence_type": None, "status": "pending",
                 "source_doc_id": "vantage-master-q2-2026", "source_page": ((n - 1) // 7) + 1}
        (BASE / "claims" / f"claim-{n:04d}.json").write_text(json.dumps(claim, indent=2) + "\n")
    (BASE / "gold.yaml").write_text(yaml.safe_dump(
        {"fixture_version": "1.0.0", "audit_id": AUDIT_ID,
         "description": "Scaled synthetic vault: 5 quarters, narrative evidence, failure-class tags.",
         "claims": gold}, sort_keys=False, allow_unicode=False))
    for sub in ("results", "review_queue"):
        (BASE / sub).mkdir(exist_ok=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--seed", type=int, default=7)
    args = parser.parse_args()
    BASE.mkdir(parents=True, exist_ok=True)
    VAULT.mkdir(exist_ok=True)
    w = build_world(args.seed)
    docs = generate_documents(w)
    claims = build_claims(w)
    make_master(claims)
    emit_metadata(claims, docs)
    from collections import Counter
    print(f"generated {AUDIT_ID}: {len(claims)} claims, {len(docs)} vault documents over {len(QUARTERS)} quarters")
    print("by status:", dict(Counter(c['status'] for c in claims)))
    print("by class: ", dict(Counter(c['failure_class'] for c in claims)))


if __name__ == "__main__":
    main()
