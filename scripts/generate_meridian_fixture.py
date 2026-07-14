"""Deterministically generate the Meridian Q1 2026 Proofbench fixture.

One source of truth emits the audit registry, 4-page master, eight vault
documents, frozen claims, gold file, and README. No randomness or LLM calls.
"""
from __future__ import annotations

import json
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path

import fitz
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
AUDIT_ID = "audit-2026-q1-meridian"
BASE = ROOT / "audits" / AUDIT_ID
MASTER = BASE / "master" / "meridian-q1-board-review.pdf"
VAULT = BASE / "vault"
NAVY, TEAL, PALE, INK, LINE = (0.09, 0.20, 0.31), (0.08, 0.47, 0.46), (0.91, 0.96, 0.95), (0.13, 0.18, 0.22), (0.78, 0.84, 0.87)

DOCS = [
    ("meridian-finance-v1-q1-2026", "finance-pack-v1.pdf", "pdf", "finance_pack"),
    ("meridian-finance-v2-q1-2026", "finance-pack-restated-v2.pdf", "pdf", "finance_pack"),
    ("meridian-operations-q1-2026", "operations-review.pdf", "pdf", "operations_review"),
    ("meridian-customer-q1-2026", "customer-metrics.xlsx", "xlsx", "customer_appendix"),
    ("meridian-regional-q1-2026", "regional-performance.pdf", "pdf", "regional_review"),
    ("meridian-treasury-q1-2026", "treasury-and-cash.pdf", "pdf", "treasury"),
    ("meridian-people-q1-2026", "people-roster.xlsx", "xlsx", "people_review"),
    ("meridian-esg-q1-2026", "sustainability-note.pdf", "pdf", "sustainability"),
]

# id, narrative, value, unit, status, source doc, page, source span, rule, notes
CLAIMS = [
 (1,"Net sales for the quarter were EUR 12.48 million.",12480000,"currency","supported",DOCS[1][0],1,"Net revenue | 12.480 | 10.960 | 11.900",{"kind":"exact"},"Vocabulary mismatch: net sales vs Net revenue; EUR millions."),
 (2,"Operating contribution reached EUR 2.14 million.",2140000,"currency","supported",DOCS[1][0],1,"Adjusted operating result | 2.140 | 1.720 | 1.960",{"kind":"exact"},"Vocabulary mismatch: operating contribution vs adjusted operating result."),
 (3,"The trading margin was 61.5%.",0.615,"percent","supported",DOCS[1][0],1,"Gross profit | 7.675 | 6.575 | 7.259",{"kind":"formula","formula":"gross_profit / net_revenue","tolerance_pct":0.2},"Formula: gross profit / net revenue = 61.5%."),
 (4,"Available liquidity at quarter end was EUR 3.62 million.",3620000,"currency","supported",DOCS[5][0],1,"Total | 3,620 | 3,180 | 3,710",{"kind":"exact"},"Hardest supported claim: liquidity binds to the first duplicate 'Total' row of the treasury ledger; source is kEUR."),
 (5,"Headcount at quarter end was 142.",142,"count","supported",DOCS[2][0],1,"FTE, period end | 142 | 136 | 130",{"kind":"exact"},"Vocabulary mismatch: headcount vs FTE."),
 (6,"The enterprise client base closed at 214.",214,"count","supported",DOCS[3][0],None,"Enterprise accounts | 214 | 201 | 188",{"kind":"exact"},"XLSX repeated-header block."),
 (7,"Recurring revenue retention was 111.2%.",1.112,"percent","supported",DOCS[3][0],None,"Net dollar retention | 111.2% | 109.8% | 108.5%",{"kind":"percent_tolerance","tolerance_pct":0.1},"Vocabulary mismatch and percentage normalization."),
 (8,"Priority requests met the response commitment in 94.2% of cases.",0.942,"percent","supported",DOCS[2][0],1,"P1 response SLA attainment | 94.2% | 92.8% | 93.0%",{"kind":"exact"},"Narrative/table vocabulary mismatch."),
 (9,"Contracted recurring value was EUR 22.60 million.",22600000,"currency","supported",DOCS[3][0],None,"Annualized subscription value | €22,600,000 | €20,940,000 | €19,870,000",{"kind":"exact"},"Currency formatting variety."),
 (10,"Average enterprise contract value was about EUR 105.6 thousand.",105607.48,"currency","supported",DOCS[3][0],None,"Enterprise accounts | 214 | 201 | 188",{"kind":"formula","formula":"annualized_subscription_value / enterprise_accounts","tolerance_pct":0.1},"Formula: EUR 22.6m / 214 customers."),
 (11,"DACH contributed EUR 5.12 million of quarterly sales.",5120000,"currency","supported",DOCS[4][0],1,"DACH | 5.120 | 4.610 | 4.820",{"kind":"exact"},"Regional table lookup."),
 (12,"UK and Ireland delivered EUR 3.44 million.",3440000,"currency","supported",DOCS[4][0],1,"UKI | 3.440 | 2.980 | 3.200",{"kind":"exact"},"Cross-document naming drift: UK and Ireland vs UKI."),
 (13,"The external delivery workforce numbered 18.",18,"count","supported",DOCS[6][0],None,"Contract personnel | 18 | 16 | 15",{"kind":"exact"},"Vocabulary mismatch in messy XLSX."),
 (14,"Location-based electricity emissions were 412 tCO2e.",412,"tCO2e","supported",DOCS[7][0],1,"Scope 2 location based | 412 | 438 | 1,690",{"kind":"exact"},"FY and quarterly periods coexist."),
 (15,"Cash burn from operations was EUR 1.284 million.",-1284000,"currency","supported",DOCS[5][0],1,"Net operating cash outflow | (1,284) | (940) | (4,230)",{"kind":"exact"},"Pure unit/sign normalization trap: kEUR and parentheses."),
 (16,"Quarterly sales per employee were about EUR 87.9 thousand.",87887.32,"currency_per_fte","supported",DOCS[1][0],1,"Net revenue | 12.480 | 10.960 | 11.900",{"kind":"formula","formula":"net_revenue / fte_period_end","tolerance_pct":0.2},"Cross-document formula: finance revenue / operations FTE."),
 (17,"There were 7.89 employees per external contractor.",7.8889,"ratio","supported",DOCS[2][0],1,"FTE, period end | 142 | 136 | 130",{"kind":"formula","formula":"fte_period_end / contract_personnel","tolerance_pct":0.2},"Cross-document formula: operations FTE / people contractors."),
 (18,"The closing workforce was 143 FTE.",143,"count","contradicted",DOCS[2][0],1,"FTE, period end | 142 | 136 | 130",{"kind":"exact"},"Subtle contradiction: right period/entity, off by one."),
 (19,"Quarterly customer attrition was 2.8%.",0.028,"percent","contradicted",DOCS[3][0],None,"Gross account attrition | 3.4% | 3.1% | 3.0%",{"kind":"exact"},"Contradicted by approved customer workbook."),
 (20,"The unresolved support queue ended at 312 cases.",312,"count","contradicted",DOCS[2][0],1,"Open case inventory | 327 | 298 | 305",{"kind":"exact"},"Contradicted; label mismatch discourages numeric-only matching."),
 (21,"New enterprise wins totalled 46.",46,"count","ambiguous",DOCS[3][0],None,"New enterprise accounts | 46 | 41 | 39",{"kind":"exact"},"Operations review says 49; equal authority, same period."),
 (22,"Qualified commercial pipeline stood at EUR 8.4 million.",8400000,"currency","ambiguous",DOCS[1][0],2,"Qualified pipeline | 8.400 | 7.920 | 8.100",{"kind":"exact"},"Regional review says EUR 8.1m; no authority ordering."),
 (23,"Formal learning time totalled 1,240 hours.",1240,"hours","ambiguous",DOCS[6][0],None,"Structured learning hours | 1,240 | 1,080 | 4,520",{"kind":"exact"},"Sustainability note says 1,190 hours."),
 # Outdated claims must assert the value as current fact -- self-referential
 # wording ("the pre-close pack reported X") is literally true and lets the
 # Verifier defensibly answer `supported`, which both variants did in the
 # first smoke test.
 (24,"Quarterly turnover was EUR 12.31 million.",12310000,"currency","outdated",DOCS[0][0],1,"Net revenue | 12.310 | 10.960 | 11.900",{"kind":"exact"},"Supported only by superseded v1; v2 restates EUR 12.48m."),
 (25,"The operating contribution came to EUR 2.05 million.",2050000,"currency","outdated",DOCS[0][0],1,"Adjusted operating result | 2.050 | 1.720 | 1.960",{"kind":"exact"},"Supported only by superseded v1; v2 reports EUR 2.14m."),
 (26,"Available liquidity stood at EUR 3.71 million.",3710000,"currency","outdated",DOCS[0][0],2,"Cash and cash equivalents | 3.710 | 3.180 | 3.500",{"kind":"exact"},"Superseded by treasury confirmation and restated pack."),
 (27,"Monthly active product seats were 6,840.",6840,"count","missing_evidence",None,None,None,{"kind":"exact"},"Similar-sounding Licensed seats (7,020) exists; must not force-match."),
 (28,"Median implementation cycle was 31 days.",31,"days","missing_evidence",None,None,None,{"kind":"exact"},"Similar value Median resolution time (31 hours) exists."),
]

DISTRACTORS = [
 ["Services revenue","1.420","1.310","1.380"],["R&D expense","(2.110)","(1.980)","(2.020)"],["Sales expense","(2.440)","(2.210)","(2.390)"],
 ["Admin expense","(1.180)","(1.090)","(1.120)"],["Deferred revenue","6.730","6.210","6.440"],["Trade receivables","4.810","4.120","4.500"],
 ["Billable consultants","63","59","56"],["Engineering FTE","48","45","43"],["Voluntary exits","6","4","5"],["Median resolution time","31","34","32"],
 ["Licensed seats","7,020","6,540","6,100"],["Expansion bookings","2.840","2.420","2.510"],["SMB accounts","486","471","455"],
 ["Partner accounts","72","68","63"],["Benelux","1.180","1.020","1.090"],["Nordics","1.360","1.240","1.480"],
 # regional rows must foot to the Total row in ALL columns (Q4 10.960, budget 11.900),
 # or the vault carries an unplanted inconsistency and the Total aggregation edge never mines
 ["France","0.890","0.810","0.850"],["Other Europe","0.490","0.300","0.460"],["Accounts payable","1,148","1,036","1,090"],
 ["Tax paid","(318)","(270)","(1,010)"],["Capital expenditure","(462)","(388)","(1,620)"],["Travel emissions","96","88","351"],
 ["Scope 1","74","69","282"],["Renewable electricity","68.0%","61.0%","64.0%"],["Women in leadership","41.0%","39.0%","40.0%"],
]

# gold_evidence for the smoke subset only (MUL-8): one claim per verdict
# status (claim-0001/0018/0021/0024/0027). Conflict/outdated claims label
# both sides; the rest of CLAIMS is unlabeled pending a full labeling pass.
GOLD_EVIDENCE = {
    1: [{"doc_id": DOCS[1][0], "source": "table", "quote": "Net revenue | 12.480 | 10.960 | 11.900"}],
    18: [{"doc_id": DOCS[2][0], "source": "table", "quote": "FTE, period end | 142 | 136 | 130"}],
    21: [
        {"doc_id": DOCS[3][0], "source": "table", "quote": "New enterprise accounts | 46 | 41 | 39"},
        {"doc_id": DOCS[2][0], "source": "table", "quote": "New enterprise accounts | 49 | 42 | 38"},
    ],
    24: [
        {"doc_id": DOCS[0][0], "source": "table", "quote": "Net revenue | 12.310 | 10.960 | 11.900"},
        {"doc_id": DOCS[1][0], "source": "table", "quote": "Net revenue | 12.480 | 10.960 | 11.900"},
    ],
    27: [],
}


def pdf_header(page, title, subtitle):
    page.insert_text((50, 48), "MERIDIAN CLOUD SYSTEMS", fontsize=9, fontname="hebo", color=TEAL)
    page.insert_text((50, 78), title, fontsize=22, fontname="hebo", color=NAVY)
    page.insert_text((50, 98), subtitle, fontsize=10, color=(0.38,0.44,0.48))
    page.draw_line((50,112),(545,112),color=LINE,width=.7)

def pdf_footer(page, n):
    page.draw_line((50,805),(545,805),color=LINE,width=.6)
    page.insert_text((50,820),"Confidential | Synthetic Proofbench fixture",fontsize=7,color=(.4,.45,.48))
    page.insert_text((510,820),f"Page {n}",fontsize=7,color=(.4,.45,.48))

def draw_table(page, y, rows, widths=None, font=7.2):
    cols=max(len(r) for r in rows); widths=widths or [495/cols]*cols; h=22
    for ri,row in enumerate(rows):
        x=50
        for ci in range(cols):
            w=widths[ci]; rect=fitz.Rect(x,y+ri*h,x+w,y+(ri+1)*h)
            page.draw_rect(rect,color=LINE,fill=NAVY if ri==0 else (PALE if ri%2==0 else (1,1,1)),width=.5)
            text=str(row[ci]) if ci<len(row) else ""
            page.insert_textbox(rect+(5,5,-4,-3),text,fontsize=font,fontname="hebo" if ri==0 else "helv",color=(1,1,1) if ri==0 else INK,align=0)
            x+=w
    return y+len(rows)*h

def save_pdf(doc,path):
    doc.set_metadata({"title":path.stem,"author":"Meridian Cloud Systems","creationDate":"D:20260415090000+02'00'","modDate":"D:20260415090000+02'00'"})
    path.parent.mkdir(parents=True,exist_ok=True); doc.save(path,garbage=4,deflate=True,clean=True,no_new_id=True)

def make_master():
    doc=fitz.open()
    groups=[CLAIMS[i:i+7] for i in range(0,28,7)]
    titles=["Financial performance","Customers and operations","Exceptions requiring review","Preliminary and unsupported metrics"]
    for pno,group in enumerate(groups,1):
        p=doc.new_page(width=595,height=842); pdf_header(p,"Q1 2026 Board Review",titles[pno-1]); y=145
        p.insert_text((50,y),f"{pno}. {titles[pno-1]}",fontsize=15,fontname="hebo",color=NAVY); y+=30
        for c in group:
            p.insert_textbox(fitz.Rect(58,y,535,y+48),f"{c[0]:02d}. {c[1]}",fontsize=10.5,lineheight=1.35,color=INK); y+=58
        pdf_footer(p,pno)
    save_pdf(doc,MASTER)

def make_pdf(path,title,pages):
    doc=fitz.open()
    for i,(subtitle,tables,notes) in enumerate(pages,1):
        p=doc.new_page(width=595,height=842); pdf_header(p,title,subtitle); y=135
        for heading,rows,widths in tables:
            p.insert_text((50,y),heading,fontsize=13,fontname="hebo",color=NAVY); y+=14
            y=draw_table(p,y,rows,widths)+22
        for note in notes:
            p.insert_textbox(fitz.Rect(50,y,545,y+55),note,fontsize=8.5,lineheight=1.25,color=INK); y+=52
        pdf_footer(p,i)
    save_pdf(doc,path)

def normalize_xlsx(path):
    tmp=path.with_suffix(".normalized.xlsx")
    with zipfile.ZipFile(path,"r") as zin, zipfile.ZipFile(tmp,"w",zipfile.ZIP_DEFLATED,compresslevel=9) as zout:
        for name in sorted(zin.namelist()):
            info=zipfile.ZipInfo(name,(1980,1,1,0,0,0)); info.compress_type=zipfile.ZIP_DEFLATED; info.external_attr=0o600<<16
            payload=zin.read(name)
            if name=="docProps/core.xml":
                payload=re.sub(rb"<dcterms:modified[^>]*>.*?</dcterms:modified>",rb'<dcterms:modified xsi:type="dcterms:W3CDTF">2026-04-15T09:00:00Z</dcterms:modified>',payload)
            zout.writestr(info,payload)
    tmp.replace(path)

def make_xlsx(path,title,blocks):
    wb=Workbook(); wb.remove(wb.active)
    fixed=datetime(2026,4,15,9,0,0)
    wb.properties.creator="Meridian Cloud Systems"; wb.properties.lastModifiedBy="Meridian Cloud Systems"
    wb.properties.created=fixed; wb.properties.modified=fixed; wb.properties.revision="1"
    for sheet_name,block_sets in blocks:
        ws=wb.create_sheet(sheet_name); ws.sheet_view.showGridLines=False
        ws.merge_cells("A1:D1"); ws["A1"]=title; ws["A1"].fill=PatternFill("solid",fgColor="17324D"); ws["A1"].font=Font(color="FFFFFF",bold=True,size=16)
        row=4
        for heading,rows in block_sets:
            ws.cell(row,1,heading).font=Font(bold=True,color="17324D",size=12); row+=1
            for values in rows:
                for col,value in enumerate(values,1): ws.cell(row,col,value)
                if values==rows[0]:
                    for cell in ws[row]: cell.fill=PatternFill("solid",fgColor="1E7A78"); cell.font=Font(color="FFFFFF",bold=True)
                row+=1
            row+=2 # repeated blocks separated by blanks
        for col,w in enumerate([32,20,20,20],1): ws.column_dimensions[get_column_letter(col)].width=w
        for row_cells in ws.iter_rows():
            for cell in row_cells: cell.alignment=Alignment(vertical="center"); cell.border=Border(bottom=Side(style="hair",color="D7E0E5"))
            label=str(row_cells[0].value or "").casefold()
            if any(word in label for word in ("retention","attrition","completion")):
                for cell in row_cells[1:4]: cell.number_format="0.0%"
            elif any(word in label for word in ("value","bookings")):
                for cell in row_cells[1:4]: cell.number_format='€#,##0'
            elif row_cells[0].value not in (None,"Metric"):
                for cell in row_cells[1:4]: cell.number_format="#,##0"
        ws.freeze_panes="A4"
    wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True
    path.parent.mkdir(parents=True,exist_ok=True); wb.save(path); normalize_xlsx(path)

def generate_documents():
    make_master()
    fin_hdr=["EUR million","Q1 2026 actual","Q4 2025 actual","Q1 2026 budget"]
    common=DISTRACTORS[:6]
    make_pdf(VAULT/DOCS[0][1],"Q1 2026 Finance Pack v1",[("Pre-close management accounts",[("Income statement",[fin_hdr,["Net revenue","12.310","10.960","11.900"],["Cost of delivery","(4.805)","(4.385)","(4.641)"],["Gross profit","7.505","6.575","7.259"],["Adjusted operating result","2.050","1.720","1.960"]]+common,[150,115,115,115])],["Superseded on 15 April 2026 by restated v2."]),("Preliminary balance sheet",[("Liquidity",[fin_hdr,["Cash and cash equivalents","3.710","3.180","3.500"],["Qualified pipeline","8.300","7.920","8.100"]],[150,115,115,115])],["PRELIMINARY - not for final reporting."])])
    make_pdf(VAULT/DOCS[1][1],"Q1 2026 Finance Pack v2",[("Restated management accounts",[("Income statement",[fin_hdr,["Net revenue","12.480","10.960","11.900"],["Cost of delivery","(4.805)","(4.385)","(4.641)"],["Gross profit","7.675","6.575","7.259"],["Adjusted operating result","2.140","1.720","1.960"]]+common,[150,115,115,115])],["Final restatement. Gross margin = Gross profit / Net revenue = 61.5%."]),("Commercial indicators",[("Pipeline",[fin_hdr,["Qualified pipeline","8.400","7.920","8.100"],["Weighted pipeline","4.620","4.180","4.430"]],[150,115,115,115])],["Approved 15 April 2026."])])
    ops=[['Metric','Q1 2026','Q4 2025','Q1 2025'],['FTE, period end','142','136','130'],['P1 response SLA attainment','94.2%','92.8%','93.0%'],['Open case inventory','327','298','305'],['New enterprise accounts','49','42','38']]+DISTRACTORS[6:11]+[
        ['Customer success FTE','21','20','18'],['Product FTE','17','16','15'],['Security incidents','3','4','5'],
        ['Platform uptime','99.96%','99.94%','99.91%'],['On-call rotations','8','8','7'],['Knowledge articles','286','251','220']]
    make_pdf(VAULT/DOCS[2][1],"Q1 2026 Operations Review",[("Workforce and service delivery",[("Operating metrics",ops,[205,95,95,95])],["All metrics are final as of 31 March 2026."])])
    make_xlsx(VAULT/DOCS[3][1],"Q1 2026 Customer Metrics",[("Metrics",[
        ("Customer base",[["Metric","Q1 2026","Q4 2025","Q1 2025"],["Enterprise accounts",214,201,188],["New enterprise accounts",46,41,39],["Gross account attrition",0.034,0.031,0.030]]),
        ("Recurring value",[["Metric","Q1 2026","Q4 2025","Q1 2025"],["Annualized subscription value",22600000,20940000,19870000],["Net dollar retention",1.112,1.098,1.085],["Licensed seats",7020,6540,6100]]),
        ("Other segments",[["Metric","Q1 2026","Q4 2025","Q1 2025"],["SMB accounts",486,471,455],["Partner accounts",72,68,63],["Expansion bookings",2840000,2420000,2510000]])
    ])])
    reg=[['Region','Q1 2026 actual','Q4 2025 actual','Q1 2026 budget'],['DACH','5.120','4.610','4.820'],['UKI','3.440','2.980','3.200']]+DISTRACTORS[14:18]+[['Total','12.480','10.960','11.900'],['Qualified commercial pipeline','8.100','7.880','8.000']]
    make_pdf(VAULT/DOCS[4][1],"Q1 2026 Regional Performance",[("Revenue by market",[("Regional scorecard",reg,[145,115,115,115])],["Regional pipeline uses local CRM qualification and is not ranked against Finance."])])
    cash=[['kEUR','31 Mar 2026','31 Dec 2025','Pre-close'],['Operating accounts','2,410','2,180','2,500'],['Deposit accounts','1,210','1,000','1,210'],['Total','3,620','3,180','3,710'],['Operating inflow','4,820','4,440','4,600'],['Supplier payments','(3,116)','(2,930)','(3,050)'],['Payroll payments','(2,988)','(2,450)','(2,730)'],['Total','(1,284)','(940)','(1,180)'],['Net operating cash outflow','(1,284)','(940)','(4,230)'],['Accounts payable','1,148','1,036','1,090'],['Near coincidence control','3,621','3,181','3,711']]
    make_pdf(VAULT/DOCS[5][1],"Q1 2026 Treasury and Cash",[("Cash movement and controls",[("Treasury ledger",cash,[185,105,105,100])],["Duplicate Total labels and interleaved subtotals are intentional. Near coincidence differs by 1 kEUR and must not mine an edge."])])
    make_xlsx(VAULT/DOCS[6][1],"Q1 2026 People Review",[("People",[
        ("Workforce",[["Metric","Q1 2026","Q4 2025","FY 2025"],["Contract personnel",18,16,15],["Permanent employees",139,133,128],["Open requisitions",12,14,11]]),
        ("Learning",[["Metric","Q1 2026","Q4 2025","FY 2025"],["Structured learning hours",1240,1080,4520],["Compliance completion",0.982,0.974,0.968],["Mentoring pairs",38,34,31]])
    ])])
    esg=[['Metric','Q1 2026','Q4 2025','FY 2025'],['Scope 2 location based','412','438','1,690'],['Structured learning hours','1,190','1,050','4,420']]+DISTRACTORS[21:]
    # Headerless >=3-column table: first row is data, intentionally misclassified by current ingest.
    headerless=[['Data centres','18','17'],['Office sites','6','6'],['Certified suppliers','43','39']]
    make_pdf(VAULT/DOCS[7][1],"Q1 2026 Sustainability Note",[("Environmental and people indicators",[("ESG scorecard",esg,[205,95,95,95]),("Asset footprint (header intentionally absent)",headerless,[260,110,110])],["Learning hours use ESG survey scope and conflict with the People Review without authority ranking."])])

def emit_metadata():
    documents=[{"doc_id":"meridian-master-q1-2026","path":str(MASTER.relative_to(ROOT)),"kind":"master","format":"pdf","tag":None}]
    documents += [{"doc_id":d,"path":str((VAULT/f).relative_to(ROOT)),"kind":"vault","format":fmt,"tag":tag} for d,f,fmt,tag in DOCS]
    audit={"audit_id":AUDIT_ID,"master_doc_id":"meridian-master-q1-2026","default_tolerance_rule":{"kind":"exact"},"evidence_priority":["finance_pack","operations_review","customer_appendix","regional_review","treasury","people_review","sustainability"],"documents":documents,"created_at":"2026-04-15T09:00:00+02:00"}
    (BASE/"audit.yaml").write_text(yaml.safe_dump(audit,sort_keys=False,allow_unicode=False))
    gold=[]
    for n,text,value,unit,status,doc,page,span,rule,notes in CLAIMS:
        entry={"claim_id":f"{AUDIT_ID}/claim-{n:04d}","claim_text":text,"canonical_value":value,"unit":unit,"currency":"EUR" if unit.startswith("currency") else None,"expected_status":status,"source_doc":doc,"source_page":page,"source_span":span,"tolerance_rule":rule,"notes":notes}
        if n in GOLD_EVIDENCE:
            entry["gold_evidence"]=GOLD_EVIDENCE[n]
        gold.append(entry)
        claim={"claim_id":f"{AUDIT_ID}/claim-{n:04d}","label":text.rstrip("."),"raw_text":text,"canonical_value":value,"unit":unit,"currency":"EUR" if unit.startswith("currency") else None,"entity":"Meridian Cloud Systems SE","time_scope":"Q1 2026","tolerance_rule":rule,"expected_evidence_type":None,"status":"pending","source_doc_id":"meridian-master-q1-2026","source_page":((n-1)//7)+1}
        (BASE/"claims").mkdir(parents=True,exist_ok=True); (BASE/"claims"/f"claim-{n:04d}.json").write_text(json.dumps(claim,indent=2)+"\n")
    (BASE/"gold.yaml").write_text(yaml.safe_dump({"fixture_version":"2.0.0","audit_id":AUDIT_ID,"description":"Deterministic Meridian Q1 2026 retrieval and verification stress fixture.","claims":gold},sort_keys=False,allow_unicode=False))
    readme=f"""# Meridian Q1 2026 audit fixture

Deterministic, fictional retrieval benchmark generated by `scripts/generate_meridian_fixture.py`. All nine documents, 28 frozen claims, and `gold.yaml` come from one source of truth. Re-running the generator is byte-identical.

## Scale and verdicts

- 4-page master; 8 vault documents (6 PDF, 2 XLSX); 28 frozen claims.
- Expected: 17 supported, 3 contradicted, 3 ambiguous, 3 outdated, 2 missing evidence.
- Verification benchmark only: extraction scoring is deferred. Frozen claim IDs map 1:1 to gold.

## Planted traps

- Vocabulary mismatch: claims 1, 2, 4-9, 13, 15.
- Cross-document naming drift: claims 7, 12, 20.
- Cross-document formulas: claims 16 and 17. Formula/tolerance cases: 3 and 10.
- Unit traps: claims 1, 4, 7, 15; Q1/Q4/FY period density throughout.
- Headerless 3-column PDF table: Sustainability Note asset footprint.
- Duplicate `Total` rows, interleaved subtotals, parentheses, mixed precision: Treasury. Claim 4 (liquidity) must bind to the first `Total` row -- the hardest supported claim.
- Regional scorecard foots to its `Total` row in all three columns, so the aggregation edge mines.
- Repeated XLSX headers with blank spacer rows: Customer and People workbooks.
- Near-coincidence: Treasury `Near coincidence control` is 1 kEUR above the component sum and must mine no edge.
- Ambiguities: 21 (46 vs 49), 22 (EUR 8.4m vs 8.1m), 23 (1,240 vs 1,190).
- Superseded v1/v2: claims 24-26. Similar-metric missing traps: claims 27-28.

## Smoke subset

One claim per verdict status: `--claims claim-0001,claim-0018,claim-0021,claim-0024,claim-0027` (about $0.35 per variant at current cost).

## Regeneration and self-test

```bash
uv run python scripts/generate_meridian_fixture.py
uv run proofbench init {AUDIT_ID}
uv run proofbench index {AUDIT_ID}
uv run python scripts/selftest_meridian_fixture.py
```

Full two-variant budget: `--max-budget-usd 4.50`. Every PDF uses real vector text and ruled tables for PyMuPDF `find_tables()` and workbench bounding boxes.
"""
    (BASE/"README.md").write_text(readme)
    for sub in ("results","review_queue"): (BASE/sub).mkdir(exist_ok=True)

def main():
    BASE.mkdir(parents=True,exist_ok=True); VAULT.mkdir(exist_ok=True)
    generate_documents(); emit_metadata()
    print(f"generated {AUDIT_ID}: {len(CLAIMS)} claims, {len(DOCS)} vault documents")

if __name__ == "__main__": main()
