"""Parse master + vault documents into index/parsed/<doc_id>.json.

Raw documents in audits/<id>/master/ and vault/ are evidence and are never
modified. Everything this module writes lives under index/ and is fully
rebuildable by re-running ingestion.
"""

from __future__ import annotations

import json
from pathlib import Path

import fitz  # PyMuPDF
import openpyxl
import yaml

from proofbench.models import AuditConfig, DocumentFormat, DocumentRef

REPO_ROOT = Path(__file__).resolve().parents[2]
INDEX_PARSED_DIR = REPO_ROOT / "index" / "parsed"


def load_audit_config(audit_id: str) -> AuditConfig:
    audit_yaml = REPO_ROOT / "audits" / audit_id / "audit.yaml"
    data = yaml.safe_load(audit_yaml.read_text())
    return AuditConfig.model_validate(data)


def parse_pdf(path: Path) -> dict:
    pages = []
    tables = []
    with fitz.open(path) as pdf:
        for page_number, page in enumerate(pdf, start=1):
            pages.append(
                {
                    "page": page_number,
                    "text": page.get_text(),
                    "width": page.rect.width,
                    "height": page.rect.height,
                }
            )
            for table_index, table in enumerate(page.find_tables().tables, start=1):
                rows = table.extract()
                if rows:
                    tables.append(
                        {
                            "page": page_number,
                            "table_index": table_index,
                            "rows": rows,
                            # one bbox per row in `rows`, same indexing (row 0 = header)
                            "row_bboxes": [list(row.bbox) for row in table.rows],
                        }
                    )
    return {"pages": pages, "tables": tables}


def parse_xlsx(path: Path) -> dict:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        cells = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cells.append({"ref": cell.coordinate, "value": _jsonable(cell.value)})
        sheets.append({"sheet": sheet_name, "cells": cells})
    return {"sheets": sheets}


def _jsonable(value):
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def parse_document(doc: DocumentRef) -> dict:
    path = REPO_ROOT / doc.path
    if doc.format == DocumentFormat.PDF:
        body = parse_pdf(path)
    elif doc.format == DocumentFormat.XLSX:
        body = parse_xlsx(path)
    else:
        raise ValueError(f"unsupported document format: {doc.format}")
    return {"doc_id": doc.doc_id, "format": doc.format.value, "source_path": doc.path, **body}


def ingest_audit(audit_id: str) -> list[Path]:
    """Parse every document registered for an audit and write index/parsed/<doc_id>.json.

    Returns the list of paths written.
    """
    config = load_audit_config(audit_id)
    INDEX_PARSED_DIR.mkdir(parents=True, exist_ok=True)
    written = []
    for doc in config.documents:
        parsed = parse_document(doc)
        out_path = INDEX_PARSED_DIR / f"{doc.doc_id}.json"
        out_path.write_text(json.dumps(parsed, indent=2) + "\n")
        written.append(out_path)
    return written
